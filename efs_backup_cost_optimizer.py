#!/usr/bin/env python3
"""
EFS & AWS Backup Cost Optimizer
Evaluates EFS file systems and AWS Backup recovery points for cost optimization opportunities.

Usage:
    python3 efs_backup_cost_optimizer.py <aws-profile> <region(s)> [max-retention-days]

Example:
    python3 efs_backup_cost_optimizer.py default us-east-1
    python3 efs_backup_cost_optimizer.py production us-east-1,eu-west-1 365
"""

import sys
import boto3
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colour fills for recommendations
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# EFS Regional pricing ($/GB-month) per region
# Source: https://aws.amazon.com/efs/pricing/
# Pricing varies by throughput mode:
#   - Elastic Throughput: Standard $0.30, IA $0.016, Archive $0.008 (us-east-1)
#   - Legacy (Bursting/Provisioned): Standard $0.30, IA $0.025, Archive $0.008 (us-east-1)
# Format: region -> { 'standard': price, 'ia_elastic': price, 'ia_legacy': price, 'archive': price }
EFS_PRICING = {
    'us-east-1': {'standard': 0.30, 'ia_elastic': 0.016, 'ia_legacy': 0.025, 'archive': 0.008},
    'us-east-2': {'standard': 0.30, 'ia_elastic': 0.016, 'ia_legacy': 0.025, 'archive': 0.008},
    'us-west-1': {'standard': 0.33, 'ia_elastic': 0.0176, 'ia_legacy': 0.0275, 'archive': 0.0088},
    'us-west-2': {'standard': 0.30, 'ia_elastic': 0.016, 'ia_legacy': 0.025, 'archive': 0.008},
    'ca-central-1': {'standard': 0.33, 'ia_elastic': 0.0176, 'ia_legacy': 0.0275, 'archive': 0.0088},
    'eu-west-1': {'standard': 0.33, 'ia_elastic': 0.0176, 'ia_legacy': 0.0275, 'archive': 0.0088},
    'eu-west-2': {'standard': 0.348, 'ia_elastic': 0.01856, 'ia_legacy': 0.029, 'archive': 0.00928},
    'eu-west-3': {'standard': 0.348, 'ia_elastic': 0.01856, 'ia_legacy': 0.029, 'archive': 0.00928},
    'eu-central-1': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'eu-central-2': {'standard': 0.396, 'ia_elastic': 0.02112, 'ia_legacy': 0.033, 'archive': 0.01056},
    'eu-north-1': {'standard': 0.312, 'ia_elastic': 0.01664, 'ia_legacy': 0.026, 'archive': 0.00832},
    'eu-south-1': {'standard': 0.348, 'ia_elastic': 0.01856, 'ia_legacy': 0.029, 'archive': 0.00928},
    'ap-southeast-1': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-southeast-2': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-southeast-3': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-northeast-1': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-northeast-2': {'standard': 0.33, 'ia_elastic': 0.0176, 'ia_legacy': 0.0275, 'archive': 0.0088},
    'ap-northeast-3': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-south-1': {'standard': 0.36, 'ia_elastic': 0.0192, 'ia_legacy': 0.03, 'archive': 0.0096},
    'ap-east-1': {'standard': 0.39, 'ia_elastic': 0.0208, 'ia_legacy': 0.0325, 'archive': 0.0104},
    'sa-east-1': {'standard': 0.45, 'ia_elastic': 0.024, 'ia_legacy': 0.0375, 'archive': 0.012},
    'me-south-1': {'standard': 0.396, 'ia_elastic': 0.02112, 'ia_legacy': 0.033, 'archive': 0.01056},
    'af-south-1': {'standard': 0.396, 'ia_elastic': 0.02112, 'ia_legacy': 0.033, 'archive': 0.01056},
}

# Fallback pricing for regions not in the map (uses us-east-1 as default)
DEFAULT_PRICING = {'standard': 0.30, 'ia_elastic': 0.016, 'ia_legacy': 0.025, 'archive': 0.008}

# AWS Backup EFS pricing ($/GB-month)
BACKUP_PRICING = {
    'warm': 0.05,
    'cold': 0.01,
}


def get_efs_pricing(region, throughput_mode):
    """Get EFS pricing for a region and throughput mode."""
    pricing = EFS_PRICING.get(region, DEFAULT_PRICING)
    # Elastic throughput uses lower IA pricing; legacy (bursting/provisioned) uses higher IA pricing
    if throughput_mode == 'elastic':
        ia_price = pricing['ia_elastic']
    else:
        ia_price = pricing['ia_legacy']
    return {
        'standard': pricing['standard'],
        'ia': ia_price,
        'archive': pricing['archive'],
    }


def get_session(profile, region):
    session = boto3.Session(profile_name=profile, region_name=region)
    return session


def evaluate_efs(session, region):
    """Evaluate EFS file systems for cost optimization."""
    efs = session.client('efs', region_name=region)
    cw = session.client('cloudwatch', region_name=region)
    backup = session.client('backup', region_name=region)
    results = []
    now = datetime.now(timezone.utc)

    # Get all EFS file systems
    paginator = efs.get_paginator('describe_file_systems')
    file_systems = []
    for page in paginator.paginate():
        file_systems.extend(page['FileSystems'])

    for fs in file_systems:
        fs_id = fs['FileSystemId']
        fs_name = next((t['Value'] for t in fs.get('Tags', []) if t['Key'] == 'Name'), fs_id)
        throughput_mode = fs.get('ThroughputMode', 'bursting')

        # Get per-storage-class sizes from SizeInBytes
        size_info = fs.get('SizeInBytes', {})
        total_size_bytes = size_info.get('Value', 0)
        standard_bytes = size_info.get('ValueInStandard', 0)
        ia_bytes = size_info.get('ValueInIA', 0)
        archive_bytes = size_info.get('ValueInArchive', 0)

        total_size_gb = round(total_size_bytes / (1024**3), 2)
        standard_gb = round(standard_bytes / (1024**3), 2)
        ia_gb = round(ia_bytes / (1024**3), 2)
        archive_gb = round(archive_bytes / (1024**3), 2)

        # Get regional pricing based on throughput mode
        prices = get_efs_pricing(region, throughput_mode)

        # Check 1: No mount targets
        mt_response = efs.describe_mount_targets(FileSystemId=fs_id)
        mount_targets = mt_response.get('MountTargets', [])
        has_mount_targets = len(mount_targets) > 0

        # Check 2: Lifecycle policy — must call describe_lifecycle_configuration
        # DescribeFileSystems does NOT return lifecycle policies
        lifecycle_policies = []
        try:
            lc_response = efs.describe_lifecycle_configuration(FileSystemId=fs_id)
            lifecycle_policies = lc_response.get('LifecyclePolicies', [])
        except Exception:
            pass

        has_lifecycle = len(lifecycle_policies) > 0
        lifecycle_config = ', '.join(
            f"{p.get('TransitionToIA', p.get('TransitionToArchive', p.get('TransitionToPrimaryStorageClass', 'N/A')))}"
            for p in lifecycle_policies
        ) if has_lifecycle else 'NONE'

        # Check 3: Size constant in last 90 days (no growth = no use)
        size_constant = False
        try:
            metrics = cw.get_metric_statistics(
                Namespace='AWS/EFS',
                MetricName='StorageBytes',
                Dimensions=[{'Name': 'FileSystemId', 'Value': fs_id}, {'Name': 'StorageClass', 'Value': 'Total'}],
                StartTime=now - timedelta(days=90),
                EndTime=now,
                Period=86400 * 30,  # monthly
                Statistics=['Average']
            )
            datapoints = sorted(metrics.get('Datapoints', []), key=lambda x: x['Timestamp'])
            if len(datapoints) >= 2:
                first = datapoints[0]['Average']
                last = datapoints[-1]['Average']
                if first > 0 and abs(last - first) / first < 0.01:  # <1% change
                    size_constant = True
        except Exception:
            pass

        # Check 4: Backup configured
        efs_backup_enabled = False
        try:
            bp = efs.describe_backup_policy(FileSystemId=fs_id)
            efs_backup_enabled = bp.get('BackupPolicy', {}).get('Status') == 'ENABLED'
        except Exception:
            pass

        aws_backup_protected = False
        try:
            selections = backup.list_protected_resources()
            for resource in selections.get('Results', []):
                if fs_id in resource.get('ResourceArn', ''):
                    aws_backup_protected = True
                    break
        except Exception:
            pass

        backup_configured = efs_backup_enabled or aws_backup_protected

        # Calculate actual monthly cost using per-storage-class sizes
        est_monthly_cost = round(
            (standard_gb * prices['standard']) +
            (ia_gb * prices['ia']) +
            (archive_gb * prices['archive']),
            2
        )

        # Determine recommendations and savings
        issues = []
        recommendation = "OK"
        fill = GREEN
        potential_savings = 0.0

        if not has_mount_targets:
            issues.append("NO MOUNT TARGETS — filesystem may be unused")
            recommendation = "REVIEW — NO MOUNT TARGETS"
            fill = RED
            potential_savings = est_monthly_cost  # Delete saves all

        if not has_lifecycle:
            issues.append("NO LIFECYCLE POLICY — missing IA transition (potential savings on Standard-class data)")
            if fill != RED:
                recommendation = "ADD LIFECYCLE POLICY"
                fill = ORANGE
            # Only calculate savings on data currently in Standard storage class
            # Assume 80% of Standard data could transition to IA
            if standard_gb > 0:
                potential_savings = round(standard_gb * 0.80 * (prices['standard'] - prices['ia']), 2)

        if size_constant:
            issues.append("SIZE CONSTANT 90 DAYS — no new data written, possible unused filesystem")
            if fill == GREEN:
                recommendation = "REVIEW — POSSIBLY UNUSED"
                fill = YELLOW

        if not backup_configured:
            issues.append("NO BACKUP — neither EFS automatic backup nor AWS Backup protection")
            if fill == GREEN:
                recommendation = "WARNING — NO BACKUP CONFIGURED"
                fill = YELLOW

        results.append({
            'region': region,
            'file_system_id': fs_id,
            'name': fs_name,
            'size_gb': total_size_gb,
            'standard_gb': standard_gb,
            'ia_gb': ia_gb,
            'archive_gb': archive_gb,
            'throughput_mode': throughput_mode,
            'has_mount_targets': has_mount_targets,
            'mount_target_count': len(mount_targets),
            'has_lifecycle': has_lifecycle,
            'lifecycle_config': lifecycle_config,
            'size_constant_90d': size_constant,
            'efs_backup_enabled': efs_backup_enabled,
            'aws_backup_protected': aws_backup_protected,
            'backup_configured': backup_configured,
            'est_monthly_cost': est_monthly_cost,
            'potential_monthly_savings': potential_savings,
            'recommendation': recommendation,
            'issues': '; '.join(issues) if issues else 'No issues found',
            'fill': fill
        })

    return results


def evaluate_backup(session, region, max_retention_days=None):
    """Evaluate AWS Backup recovery points for cost optimization."""
    backup = session.client('backup', region_name=region)
    results = []
    now = datetime.now(timezone.utc)

    # Get all backup vaults
    vaults = []
    paginator = backup.get_paginator('list_backup_vaults')
    for page in paginator.paginate():
        vaults.extend(page['BackupVaultList'])

    for vault in vaults:
        vault_name = vault['BackupVaultName']

        # Get recovery points
        rp_paginator = backup.get_paginator('list_recovery_points_by_backup_vault')
        for page in rp_paginator.paginate(BackupVaultName=vault_name):
            for rp in page['RecoveryPoints']:
                rp_arn = rp['RecoveryPointArn']
                resource_type = rp.get('ResourceType', 'Unknown')
                resource_arn = rp.get('ResourceArn', '')
                creation_date = rp.get('CreationDate', now)
                status = rp.get('Status', 'Unknown')
                lifecycle = rp.get('Lifecycle', {})
                delete_after = lifecycle.get('DeleteAfterDays')
                move_to_cold = lifecycle.get('MoveToColdStorageAfterDays')
                backup_size_bytes = rp.get('BackupSizeInBytes', 0)
                backup_size_gb = round(backup_size_bytes / (1024**3), 2) if backup_size_bytes else 0

                age_days = (now - creation_date).days if creation_date else 0

                # Determine expiration
                if delete_after:
                    expiry_date = creation_date + timedelta(days=delete_after)
                    days_to_expiry = (expiry_date - now).days
                    retention_type = f"{delete_after} days"
                else:
                    expiry_date = None
                    days_to_expiry = None
                    retention_type = "NEVER (infinite)"

                # Storage tier based on lifecycle setting and age
                storage_tier = "Warm"
                if move_to_cold and age_days > move_to_cold:
                    storage_tier = "Cold"

                issues = []
                recommendation = "OK"
                fill = GREEN

                # Check 1: Infinite retention
                if delete_after is None:
                    issues.append(f"INFINITE RETENTION — recovery point will never expire ({age_days} days old)")
                    recommendation = "REVIEW — SET RETENTION POLICY"
                    fill = ORANGE

                # Check 2: Exceeds max retention
                if max_retention_days and age_days > max_retention_days:
                    issues.append(f"EXCEEDS MAX RETENTION — {age_days} days old (max: {max_retention_days})")
                    recommendation = "DELETE — EXCEEDS RETENTION"
                    fill = YELLOW

                # Check 3: Expired status
                if status == 'EXPIRED':
                    issues.append("EXPIRED STATUS — recovery point marked as expired")
                    recommendation = "DELETE — EXPIRED"
                    fill = RED

                # Estimate cost
                if storage_tier == "Cold":
                    est_monthly_cost = round(backup_size_gb * BACKUP_PRICING['cold'], 2)
                else:
                    est_monthly_cost = round(backup_size_gb * BACKUP_PRICING['warm'], 2)

                results.append({
                    'region': region,
                    'vault_name': vault_name,
                    'recovery_point_arn': rp_arn,
                    'resource_type': resource_type,
                    'resource_arn': resource_arn,
                    'status': status,
                    'storage_tier': storage_tier,
                    'creation_date': creation_date.strftime('%Y-%m-%d') if creation_date else '',
                    'age_days': age_days,
                    'retention_type': retention_type,
                    'delete_after_days': delete_after or 'NEVER',
                    'move_to_cold_days': move_to_cold or 'N/A',
                    'days_to_expiry': days_to_expiry if days_to_expiry is not None else 'NEVER',
                    'backup_size_gb': backup_size_gb,
                    'est_monthly_cost': est_monthly_cost,
                    'recommendation': recommendation,
                    'issues': '; '.join(issues) if issues else 'No issues found',
                    'fill': fill
                })

    return results


def calculate_savings(efs_results, backup_results):
    """Calculate total potential savings."""
    # EFS lifecycle savings (only on Standard-class data for filesystems without lifecycle)
    efs_lifecycle_savings = sum(
        r['potential_monthly_savings'] for r in efs_results
        if not r['has_lifecycle'] and r['has_mount_targets']
    )
    # EFS unused savings (no mount targets = delete candidate)
    efs_unused_savings = sum(
        r['est_monthly_cost'] for r in efs_results if not r['has_mount_targets']
    )

    backup_savings_infinite = sum(r['est_monthly_cost'] for r in backup_results if r['retention_type'] == 'NEVER (infinite)')
    backup_savings_expired = sum(r['est_monthly_cost'] for r in backup_results if r['status'] == 'EXPIRED')
    backup_savings_exceeded = sum(r['est_monthly_cost'] for r in backup_results if 'EXCEEDS' in r['recommendation'])

    total_monthly = efs_lifecycle_savings + efs_unused_savings + backup_savings_expired + backup_savings_exceeded

    return {
        'efs_lifecycle_savings': efs_lifecycle_savings,
        'efs_unused_savings': efs_unused_savings,
        'backup_infinite_retention': backup_savings_infinite,
        'backup_expired': backup_savings_expired,
        'backup_exceeded_retention': backup_savings_exceeded,
        'total_monthly': total_monthly,
        'total_annual': total_monthly * 12,
    }


def write_excel(efs_results, backup_results, savings, regions, max_retention):
    """Generate colour-coded Excel report."""
    wb = Workbook()

    # Sheet 1: EFS Assessment
    ws_efs = wb.active
    ws_efs.title = "EFS Cost Optimization"
    efs_headers = ['Region', 'File System ID', 'Name', 'Total Size (GB)', 'Standard (GB)',
                   'IA (GB)', 'Archive (GB)', 'Throughput Mode', 'Mount Targets',
                   'Has Lifecycle', 'Lifecycle Config', 'Size Constant 90d',
                   'Backup Configured', 'Est Monthly Cost ($)',
                   'Potential Savings ($/mo)', 'Recommendation', 'Issues']
    for col, h in enumerate(efs_headers, 1):
        cell = ws_efs.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for row_idx, r in enumerate(efs_results, 2):
        values = [r['region'], r['file_system_id'], r['name'], r['size_gb'],
                  r['standard_gb'], r['ia_gb'], r['archive_gb'], r['throughput_mode'],
                  r['mount_target_count'], r['has_lifecycle'], r['lifecycle_config'],
                  r['size_constant_90d'], r['backup_configured'],
                  r['est_monthly_cost'], r['potential_monthly_savings'],
                  r['recommendation'], r['issues']]
        for col, val in enumerate(values, 1):
            cell = ws_efs.cell(row=row_idx, column=col, value=val)
            cell.fill = r['fill']
            cell.border = THIN_BORDER

    for col in range(1, len(efs_headers) + 1):
        ws_efs.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Backup Assessment
    ws_backup = wb.create_sheet("AWS Backup Optimization")
    backup_headers = ['Region', 'Vault', 'Resource Type', 'Resource ARN', 'Status', 'Storage Tier',
                      'Creation Date', 'Age (Days)', 'Retention', 'Move to Cold (Days)',
                      'Days to Expiry', 'Size (GB)', 'Est Monthly Cost ($)',
                      'Recommendation', 'Issues']
    for col, h in enumerate(backup_headers, 1):
        cell = ws_backup.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for row_idx, r in enumerate(backup_results, 2):
        values = [r['region'], r['vault_name'], r['resource_type'], r['resource_arn'], r['status'],
                  r['storage_tier'], r['creation_date'], r['age_days'], r['retention_type'],
                  r['move_to_cold_days'], r['days_to_expiry'],
                  r['backup_size_gb'], r['est_monthly_cost'], r['recommendation'], r['issues']]
        for col, val in enumerate(values, 1):
            cell = ws_backup.cell(row=row_idx, column=col, value=val)
            cell.fill = r['fill']
            cell.border = THIN_BORDER

    for col in range(1, len(backup_headers) + 1):
        ws_backup.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 3: Cost Savings Summary
    ws_savings = wb.create_sheet("Cost Savings Summary")
    ws_savings.column_dimensions['A'].width = 45
    ws_savings.column_dimensions['B'].width = 25

    summary_data = [
        ("Cost Savings Summary", ""),
        ("", ""),
        ("Category", "Monthly Savings ($)"),
        ("EFS — Add Lifecycle Policy (IA transition on Standard data)", f"${savings['efs_lifecycle_savings']:.2f}"),
        ("EFS — Delete Unused Filesystems (no mount targets)", f"${savings['efs_unused_savings']:.2f}"),
        ("Backup — Delete Expired Recovery Points", f"${savings['backup_expired']:.2f}"),
        ("Backup — Delete Over-Retained Recovery Points", f"${savings['backup_exceeded_retention']:.2f}"),
        ("Backup — Infinite Retention (review needed)", f"${savings['backup_infinite_retention']:.2f}"),
        ("", ""),
        ("TOTAL MONTHLY SAVINGS", f"${savings['total_monthly']:.2f}"),
        ("TOTAL ANNUAL SAVINGS", f"${savings['total_annual']:.2f}"),
        ("", ""),
        ("Parameters", ""),
        ("Regions", ', '.join(regions)),
        ("Max Retention (days)", str(max_retention) if max_retention else "Not specified"),
        ("Run Date", datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')),
        ("", ""),
        ("Notes", ""),
        ("", "EFS lifecycle savings are calculated only on data currently in Standard storage class."),
        ("", "Cost estimates use region-specific pricing and per-filesystem throughput mode."),
        ("", "Savings assume 80% of Standard-class data would transition to IA with a lifecycle policy."),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_savings.cell(row=row_idx, column=1, value=label).font = Font(bold=(row_idx in [1, 3, 10, 11, 13]))
        ws_savings.cell(row=row_idx, column=2, value=value)

    # Save
    filename = f"efs_backup_cost_optimization_{'_'.join(regions)}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 efs_backup_cost_optimizer.py <aws-profile> <region(s)> [max-retention-days]")
        print("Example: python3 efs_backup_cost_optimizer.py default us-east-1,eu-west-1 365")
        sys.exit(1)

    profile = sys.argv[1]
    regions = sys.argv[2].split(',')
    max_retention = int(sys.argv[3]) if len(sys.argv) > 3 else None

    print(f"╔══════════════════════════════════════════════════╗")
    print(f"║   EFS & AWS Backup Cost Optimizer               ║")
    print(f"╠══════════════════════════════════════════════════╣")
    print(f"║  Profile: {profile:<38} ║")
    print(f"║  Regions: {', '.join(regions):<38} ║")
    print(f"║  Max Retention: {str(max_retention) + ' days' if max_retention else 'Not set':<32} ║")
    print(f"╚══════════════════════════════════════════════════╝")
    print()

    all_efs_results = []
    all_backup_results = []

    for region in regions:
        region = region.strip()
        print(f"[{region}] Connecting...")
        session = get_session(profile, region)

        print(f"[{region}] Evaluating EFS file systems...")
        efs_results = evaluate_efs(session, region)
        all_efs_results.extend(efs_results)
        print(f"[{region}]   → {len(efs_results)} file systems evaluated")

        print(f"[{region}] Evaluating AWS Backup recovery points...")
        backup_results = evaluate_backup(session, region, max_retention)
        all_backup_results.extend(backup_results)
        print(f"[{region}]   → {len(backup_results)} recovery points evaluated")
        print()

    # Calculate savings
    savings = calculate_savings(all_efs_results, all_backup_results)

    # Generate report
    print("Generating Excel report...")
    filename = write_excel(all_efs_results, all_backup_results, savings, regions, max_retention)
    print(f"✅ Report saved: {filename}")
    print()
    print(f"Summary:")
    print(f"  EFS file systems evaluated: {len(all_efs_results)}")
    print(f"  Backup recovery points evaluated: {len(all_backup_results)}")
    print(f"  Estimated monthly savings: ${savings['total_monthly']:.2f}")
    print(f"  Estimated annual savings:  ${savings['total_annual']:.2f}")


if __name__ == '__main__':
    main()
